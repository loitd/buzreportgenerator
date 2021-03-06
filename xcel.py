from openpyxl import Workbook, load_workbook
from halo import Halo
import time
from datetime import date, timedelta, datetime
from copy import copy
class XCel:
    def __init__(self, config):
        p1 = time.time()
        self.config = config
        spin = Halo(text="Opening template file: {0}".format(self.config['SALE_REPORT_TEMPLATE']), spinner='dots')
        spin.start()
        self.wb = load_workbook(self.config['SALE_REPORT_TEMPLATE'], keep_vba=True)
        self.ws = self.wb[self.config['SHEET_DATA_RAW']]
        self.ws2 = self.wb[self.config['SHEET_MERCHANT']]
        p2 = time.time()
        spin.stop()
        print("[{0}] Excel file opened!".format(p2-p1))
    
    def __del__(self):
        self.wb.close()
    
    def save(self, filename="./.data/Summary-sale-report-{0}.xls"):
        p1 = time.time()
        spin = Halo(text="Begin saving file to: {0}".format(filename), spinner = 'dots')
        spin.start()
        # For more date format:https://docs.python.org/2/library/datetime.html#strftime-and-strptime-behavior
        yesterday = date.today() - timedelta(1)
        self.wb.save(filename=filename.format(yesterday.strftime("%Y%m%d")))
        p2 = time.time()
        spin.stop()
        print("[{0}] Excel file saved!".format(p2-p1))
    
    def detectBlank(self, col=1, ws=1):
        i = 1
        if ws == 1:
            wsx = self.ws
        elif ws == 2:
            wsx = self.ws2

        while 1:
            if wsx.cell(column=col, row=i).value is None:
                return i
                break
            else:
                i = i + 1

    # Begin appending datas into first blank cell of the col
    # By default, append to ws1 = Data raw
    def appendNext(self, datas, col, ws=1, firstColDatetime=True):
        p1 = time.time()
        print("Begin appendNext with firstCol={0}".format(col))
        # First, detect for first bank row
        row = self.detectBlank(ws=ws, col=col)
        oriRow = row
        oriCol = col
        # check ws
        if ws == 1:
            wsx = self.ws
        elif ws == 2:
            wsx = self.ws2
        # get one sample format of 1st col
        cellx = wsx.cell(column=col, row=row-1)
        # Now writting, datas is a list, data is tuple
        for data in datas:
            print("Writing: {0} at row/col: {1}/{2}".format(data, row, col))
            # data is a tuple, dat is an element
            for dat in data:
                if firstColDatetime == True and col == oriCol:
                    datx = datetime.strptime(dat,"%m/%d/%Y")
                else:
                    datx = dat
                wsx.cell(column=col, row=row, value=datx)
                col = col + 1
            row = row + 1
            col = oriCol
        # Now copy format for first col
        print("Begin formatting rows ...")
        for rows in wsx.iter_rows(min_row=oriRow, max_row=row, min_col=oriCol, max_col=oriCol):
            for cell in rows:
                if cellx.has_style:
                    cell.number_format = copy(cellx.number_format)
                    cell.alignment = copy(cellx.alignment)
        p2 = time.time()
        print("[{0}] appendNext complete successfully".format(p2-p1))
        return True
    
    def appendMegabank(self, datas, firstColDatetime=True):
        print("Begin append data for Megabank")
        return self.appendNext(datas, col=1)
    
    def appendTopup(self, datas, firstColDatetime=True):
        print("Begin append data for Topup")
        return self.appendNext(datas, col=16)
    
    def appendSQLTopup(self, datas, firstColDatetime=True):
        print("Begin append data for SQL Topup")
        return self.appendNext(datas, col=16)
    
    def appendSAT(self, datas, firstColDatetime=True):
        print("Begin append data for SAT")
        return self.appendNext(datas, col=24)
    
    def appendMC(self, datas, firstColDatetime=True):
        print("Begin append data for MC")
        return self.appendNext(datas, col=51)
    
    def appendVA(self, datas, firstColDatetime=True):
        print("Begin append data for VA")
        return self.appendNext(datas, col=59)
    
    def appendMD(self, datas, firstColDatetime=True):
        print("Begin append data for MD")
        return self.appendNext(datas, col=67)
    
    def appendFD(self, datas, firstColDatetime=True):
        print("Begin append data for FD")
        return self.appendNext(datas, col=75)
    
    def appendVERIFY(self, datas, firstColDatetime=True):
        print("Begin append data for VERIFY")
        return self.appendNext(datas, col=82)
    
    def appendTHS(self, datas, firstColDatetime=True):
        print("Begin append data for THS")
        return self.appendNext(datas, col=88)
    
    def appendNewMerchant(self, datas):
        print("Begin append data for New Merchant")
        return self.appendNext(datas, col=1, ws=2)

# Write something
# ws.cell(column=2, row=3, value="ahihi")
if __name__ == "__main__":
    pass

